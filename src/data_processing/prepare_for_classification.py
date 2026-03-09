import laspy
import numpy as np
import pathlib as pth
import torch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

import sys
src_path = pth.Path(__file__).parent.parent
sys.path.append(str(src_path))

from utils.visualize_trees import save_tree_projections_pdf
from array_processing_RE import TreeSegmRay


def create_species_workbook(species_dict: dict) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()

    header_font  = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", start_color="2E7D32")
    center       = Alignment(horizontal="center")
    normal_font  = Font(name="Arial")

    # ── Sheet 1: Tree Labels ──────────────────────────────────────────────────
    ws_data = wb.active
    ws_data.title = "Tree Labels"

    for col, h in enumerate(["File Name", "Tree Label"], 1):
        cell = ws_data.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center

    ws_data.column_dimensions["A"].width = 30
    ws_data.column_dimensions["B"].width = 14
    ws_data.freeze_panes = "A2"   # keep header row visible while scrolling

    # ── Sheet 2: Species Reference ───────────────────────────────────────────
    ws_ref = wb.create_sheet(title="Species Reference")

    for col, h in enumerate(["Species Code", "Species (Latin)", "Species (Polish)"], 1):
        cell = ws_ref.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center

    for row_idx, (code, names) in enumerate(species_dict.items(), 2):
        ws_ref.cell(row=row_idx, column=1, value=code).font   = normal_font
        ws_ref.cell(row=row_idx, column=2, value=names[0]).font = normal_font
        ws_ref.cell(row=row_idx, column=3, value=names[1]).font = normal_font

    ws_ref.column_dimensions["A"].width = 14
    ws_ref.column_dimensions["B"].width = 28
    ws_ref.column_dimensions["C"].width = 28
    ws_ref.freeze_panes = "A2"

    return wb


def append_tree_rows(ws, las_idx: int, tree_labels: np.ndarray, data_row_start: int) -> int:
    unique_labels = sorted(set(tree_labels[tree_labels != -1].tolist()))
    for tree_id in unique_labels:
        npy_name = f"{las_idx}_{tree_id}.npy"
        ws.cell(row=data_row_start, column=1, value=npy_name).font = Font(name="Arial")
        # column 2 (Tree Label) intentionally left empty for manual annotation
        data_row_start += 1
    return data_row_start


def main(species_dict: dict = None):

    species = {
        0:  ["Betula_pendula",         "Brzoza brodawkowata"],
        1:  ["Fagus_sylvatica",        "Buk zwyczajny"],
        2:  ["Quercus_petraea",        "Dąb bezszypułkowy"],
        3:  ["Quercus_rubra",          "Dąb czerwony"],
        4:  ["Quercus_robur",          "Dąb szypułkowy"],
        5:  ["Carpinus_betulus",       "Grab pospolity"],
        6:  ["Fraxinus_excelsior",     "Jesion wyniosły"],
        7:  ["Acer_pseudoplatanus",    "Klon jawor"],
        8:  ["Acer_campestre",         "Klon polny"],
        9:  ["Tilia_cordata",          "Lipa drobnolistna"],
        10: ["Ulmus_laevis",           "Wiąz szypułkowy"],
        11: ["Crataegus_monogyna",     "Głóg jednoszyjkowy"],
        12: ["Corylus_avellana",       "Leszczyna pospolita"],
        13: ["Pseudotsuga_menziesii",  "Daglezja zielona"],
        14: ["Abies_alba",             "Jodła pospolita"],
        15: ["Larix_decidua",          "Modrzew europejski"],
        16: ["Pinus_sylvestris",       "Sosna zwyczajna"],
        17: ["Picea_abies",            "Świerk pospolity"],
        18: ["Other",                  "Inne"],
        19: ["Incorrect segmentation", "Błędna segmentacja"],
    }

    semantic_labelled_dir = pth.Path("data/split/")
    laz_paths = list(semantic_labelled_dir.glob("*.laz"))

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    seg = TreeSegmRay(
        height_min=0.9, max_diameter=0.8, distance_limit=0.25,
        gravity_factor=0.6, use_rays=False, ground_label=1,
        tree_label=7, verbose=False,
    )
    seg.start_container()

    for las_idx, path in enumerate(laz_paths):
        dataset_dir = path.parent / path.stem
        dataset_dir.mkdir(exist_ok=True, parents=True)

        try:
            las = laspy.read(path)
            pts = np.vstack([las.x, las.y, las.z]).T
            cls = np.asarray(las.classification, dtype=np.int32)

            tree_labels = seg.segment(pts, cls)
            tree_xyz    = pts[cls == seg.tree_label]

            save_tree_projections_pdf(
                points=tree_xyz[tree_labels != -1],
                labels=tree_labels[tree_labels != -1],
                las_idx=las_idx,
                source_name=path.stem,
                output_name=f"{path.stem}.pdf",
                output_dir=dataset_dir,
                resolution=350,
                device=device,
            )

            if species_dict is not None:
                wb = create_species_workbook(species_dict)
                append_tree_rows(wb["Tree Labels"], las_idx, tree_labels, data_row_start=2)
                excel_path = dataset_dir / f"{path.stem}.xlsx"
                wb.save(excel_path)

        finally:
            seg.rm_container()


if __name__ == "__main__":
    main(species_dict = {
        0:  ["Betula_pendula",         "Brzoza brodawkowata"],
        1:  ["Fagus_sylvatica",        "Buk zwyczajny"],
        2:  ["Quercus_petraea",        "Dąb bezszypułkowy"],
        3:  ["Quercus_rubra",          "Dąb czerwony"],
        4:  ["Quercus_robur",          "Dąb szypułkowy"],
        5:  ["Carpinus_betulus",       "Grab pospolity"],
        6:  ["Fraxinus_excelsior",     "Jesion wyniosły"],
        7:  ["Acer_pseudoplatanus",    "Klon jawor"],
        8:  ["Acer_campestre",         "Klon polny"],
        9:  ["Tilia_cordata",          "Lipa drobnolistna"],
        10: ["Ulmus_laevis",           "Wiąz szypułkowy"],
        11: ["Crataegus_monogyna",     "Głóg jednoszyjkowy"],
        12: ["Corylus_avellana",       "Leszczyna pospolita"],
        13: ["Pseudotsuga_menziesii",  "Daglezja zielona"],
        14: ["Abies_alba",             "Jodła pospolita"],
        15: ["Larix_decidua",          "Modrzew europejski"],
        16: ["Pinus_sylvestris",       "Sosna zwyczajna"],
        17: ["Picea_abies",            "Świerk pospolity"],
        18: ["Other",                  "Inne"],
        19: ["Incorrect segmentation", "Błędna segmentacja"],
    }
)