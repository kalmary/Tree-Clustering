"""
reclassify_trees.py
===================
Post-processing pass over already-segmented tree data.

Source layout (one sub-dir per .laz file):
    src_root/
    └── BRIK_Grajewo_2026_6_8_mod/
        ├── BRIK_Grajewo_2026_6_8_mod.pdf
        ├── BRIK_Grajewo_2026_6_8_mod.xlsx
        └── clouds/
            ├── 51_0.npy
            ├── 51_100.npy
            └── ...

Goal layout (mirrors source; clouds renamed with label suffix):
    dst_root/
    └── BRIK_Grajewo_2026_6_8_mod/
        ├── BRIK_Grajewo_2026_6_8_mod.pdf      ← copied verbatim
        ├── BRIK_Grajewo_2026_6_8_mod.xlsx      ← col C filled
        └── clouds/
            ├── 51_0_3.npy                      ← _<label> appended
            ├── 51_100_17.npy
            └── ...
"""

from __future__ import annotations

import pathlib
import shutil
from typing import Optional

import numpy as np
import openpyxl
from openpyxl.styles import Font
from tqdm import tqdm

from GPT_TreeClassifier import LLM_Classifier


# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────

def _npy_stem_to_row(ws, stem: str) -> Optional[int]:
    """Return the 1-based row index in *ws* whose column-A value matches *stem*.npy.

    Column A stores filenames like ``51_102.npy``.
    ``stem`` is the filename without extension, e.g. ``51_102``.
    Returns None if not found.
    """
    target = f"{stem}.npy"
    for row in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=1).value
        if cell_val is not None and str(cell_val).strip() == target:
            return row
    return None


def _label_code_from_prediction(prediction: str, species_dict: dict) -> Optional[int]:
    """Map a raw prediction string back to its integer species code.

    Tries to match against both the Latin name (index 0) and the Polish name
    (index 1) stored in *species_dict*.  Returns None on no match.
    """
    pred_clean = prediction.strip()
    for code, (latin, polish) in species_dict.items():
        if pred_clean in (latin, polish, str(code)):
            return code
    return None


# ──────────────────────────────────────────────────────────────────────────────
# Core function
# ──────────────────────────────────────────────────────────────────────────────

def reclassify_trees(
    src_root: pathlib.Path | str,
    dst_root: pathlib.Path | str,
    classifier: LLM_Classifier,
    species_dict: dict,
    overwrite: bool = False,
) -> None:
    """Classify pre-segmented tree clouds and write results to a new directory.

    Parameters
    ----------
    src_root:
        Root directory containing one sub-folder per original .laz file.
    dst_root:
        Destination root.  Mirrors the source layout exactly; created if absent.
    classifier:
        Instantiated ``LLM_Classifier`` (``predict`` method must accept an
        (N, 3) float32/float64 ndarray and return a species string).
    species_dict:
        Same ``SPECIES`` dict used during segmentation — maps int code → [Latin, Polish].
    overwrite:
        When True, existing files in *dst_root* are silently overwritten.
        When False (default), sub-directories that already contain an .xlsx
        with non-empty column C are skipped.
    """
    src_root = pathlib.Path(src_root)
    dst_root = pathlib.Path(dst_root)
    dst_root.mkdir(parents=True, exist_ok=True)

    # Each immediate child of src_root that contains a clouds/ sub-dir is a batch.
    batch_dirs = sorted(
        d for d in src_root.iterdir()
        if d.is_dir() and (d / "clouds").is_dir()
    )

    if not batch_dirs:
        print(f"[RECLASSIFY] No batch directories found under {src_root}")
        return

    print(f"[RECLASSIFY] {len(batch_dirs)} batch(es) found")

    for batch_src in tqdm(batch_dirs, desc="Batches"):
        batch_dst = dst_root / batch_src.name
        clouds_dst = batch_dst / "clouds"
        clouds_dst.mkdir(parents=True, exist_ok=True)

        # ── Locate source xlsx ────────────────────────────────────────────────
        xlsx_candidates = list(batch_src.glob("*.xlsx"))
        if not xlsx_candidates:
            print(f"  [SKIP] No .xlsx in {batch_src.name}")
            continue
        xlsx_src = xlsx_candidates[0]
        xlsx_dst = batch_dst / xlsx_src.name

        # ── Skip check ───────────────────────────────────────────────────────
        if not overwrite and xlsx_dst.exists():
            wb_check = openpyxl.load_workbook(xlsx_dst)
            ws_check = wb_check["Tree Labels"]
            # If column C row 2 is already filled, assume batch is done.
            if ws_check.cell(row=2, column=3).value not in (None, ""):
                print(f"  [SKIP] {batch_src.name} already classified")
                continue

        # ── Copy PDF verbatim ────────────────────────────────────────────────
        for pdf_src in batch_src.glob("*.pdf"):
            pdf_dst = batch_dst / pdf_src.name
            if overwrite or not pdf_dst.exists():
                shutil.copy2(pdf_src, pdf_dst)

        # ── Load xlsx into memory ────────────────────────────────────────────
        wb = openpyxl.load_workbook(xlsx_src)
        ws = wb["Tree Labels"]
        green_font = Font(name="Arial", color="008000")  # verified → green

        # ── Iterate over cloud files ─────────────────────────────────────────
        clouds_src = batch_src / "clouds"
        npy_files  = sorted(clouds_src.glob("*.npy"))

        for npy_src in tqdm(npy_files, desc=f"  {batch_src.name}", leave=False):
            stem = npy_src.stem  # e.g. "51_102"

            # Load point cloud
            try:
                pts = np.load(npy_src).astype(np.float32)
            except Exception as e:
                print(f"    [WARN] Could not load {npy_src.name}: {e}")
                continue

            # Run classifier
            try:
                prediction_str = str(classifier.predict(pts))
            except Exception as e:
                print(f"    [WARN] Classifier failed on {npy_src.name}: {e}")
                prediction_str = ""

            # Resolve integer label code from prediction string
            label_code = _label_code_from_prediction(prediction_str, species_dict)
            label_suffix = str(label_code) if label_code is not None else "unk"

            # Copy .npy with renamed filename
            dst_name = f"{stem}_{label_suffix}.npy"
            npy_dst  = clouds_dst / dst_name
            if overwrite or not npy_dst.exists():
                shutil.copy2(npy_src, npy_dst)

            # Write prediction to Excel column C
            row = _npy_stem_to_row(ws, stem)
            if row is not None:
                cell = ws.cell(row=row, column=3, value=prediction_str)
                cell.font = green_font
            else:
                print(f"    [WARN] No Excel row found for {stem}.npy")

        # ── Save updated xlsx ────────────────────────────────────────────────
        wb.save(xlsx_dst)
        print(f"  [OK] {batch_src.name} → {xlsx_dst.relative_to(dst_root)}")

    print(f"[RECLASSIFY] Done.  Results in {dst_root.resolve()}")


# ──────────────────────────────────────────────────────────────────────────────
# Entry point
# ──────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import os
    from dotenv import load_dotenv

    load_dotenv()

    SPECIES = {
        0:  ["Betula_pendula",         "Brzoza brodawkowata"],
        1:  ["Fagus_sylvatica",        "Buk zwyczajny"],
        2:  ["Quercus_petraea",        "Dąb bezszypułkowy"],
        3:  ["Quercus_rubra",          "Dąb czerwony"],
        4:  ["Quercus_robur",          "Dąb szypułkowy"],
        5:  ["Carpinus_betulus",       "Grab pospolity"],
        6:  ["Fraxinus_excelsior",     "Jesion wyniosły"],
        7:  ["Acer_pseudoplatanus",    "Klon jawor"],
        8:  ["Acer_campestre",         "Klon polny"],
        9:  ["Tilia_cordata",          "Lipa drobnolistna"],
        10: ["Ulmus_laevis",           "Wiąz szypułkowy"],
        11: ["Crataegus_monogyna",     "Głóg jednoszyjkowy"],
        12: ["Corylus_avellana",       "Leszczyna pospolita"],
        13: ["Pseudotsuga_menziesii",  "Daglezja zielona"],
        14: ["Abies_alba",             "Jodła pospolita"],
        15: ["Larix_decidua",          "Modrzew europejski"],
        16: ["Pinus_sylvestris",       "Sosna zwyczajna"],
        17: ["Picea_abies",            "Świerk pospolity"],
        18: ["Other",                  "Inne"],
        19: ["Incorrect segmentation", "Błędna segmentacja"],
    }
    import os
    from dotenv import load_dotenv

    load_dotenv()

    OPENAI_API_KEY = os.getenv('OPENAI_API_KEY')
    MODEL = "gpt-5.4"
    # from LLM_TreeClassifier import LLM_Classifier

    classifier = LLM_Classifier(resolution=350,
                                species=SPECIES,
                                API_KEY=OPENAI_API_KEY,
                                model=MODEL)

    reclassify_trees(
        src_root    = pathlib.Path("/mnt/DATA_SSD/BRIK/GRAJEWO_CUT"),
        dst_root    = pathlib.Path("/mnt/DATA_SSD/BRIK/GRAJEWO_CUT2"),
        classifier  = classifier,
        species_dict= SPECIES,
        overwrite   = True,
    )