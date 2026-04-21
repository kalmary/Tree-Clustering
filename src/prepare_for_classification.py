import laspy
import numpy as np
import pathlib as pth
import torch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from utils.visualize_trees import save_tree_projections_pdf
from utils.plot_cloud import plot_cloud
from array_processing_RE import TreeSegmRay

<<<<<<< HEAD
from api_multiple_classifier import LLM_Classifier
=======
from GPT_TreeClassifier import LLM_Classifier
>>>>>>> main
from tqdm import tqdm

import os
from dotenv import load_dotenv

load_dotenv()

OPENAI_API_KEY = os.getenv('OPENAI_API_KEY')
MODEL = "gpt-5.4"


# ──────────────────────────────────────────────────────────────────────────────
# Label translation  (old model  →  new SPECIES dict)
# ──────────────────────────────────────────────────────────────────────────────

# Old model classes ordered alphabetically as LabelEncoder produces them,
# with "others" appended last (index = n_species).  The old model has no
# "Incorrect segmentation" concept — that label (key 19 in SPECIES) is
# reserved exclusively for manual annotation by a specialist.
#
# Update this list if you retrain the old model with different species.
def _header_cell(cell, text: str):
    cell.value     = text
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF")
    cell.fill      = PatternFill("solid", start_color="2E7D32")
    cell.alignment = Alignment(horizontal="center")


def create_workbook(species_dict: dict) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()

    # ── Sheet 1: Tree Labels ──────────────────────────────────────────────────
    ws_labels = wb.active
    ws_labels.title = "Tree Labels"

    _header_cell(ws_labels.cell(row=1, column=1), "File Name")
    _header_cell(ws_labels.cell(row=1, column=2), "Tree Label (auto)")
    _header_cell(ws_labels.cell(row=1, column=3), "Tree Label (verified)")

    ws_labels.freeze_panes = "A2"
    ws_labels.column_dimensions["A"].width = 30
    ws_labels.column_dimensions["B"].width = 26
    ws_labels.column_dimensions["C"].width = 26

    # ── Sheet 2: Species lookup ───────────────────────────────────────────────
    ws_species = wb.create_sheet("Available labels lookup")

    for col, h in enumerate(["Species Code", "Species (Latin)", "Species (Polish)"], 1):
        _header_cell(ws_species.cell(row=1, column=col), h)

    for row_idx, (code, names) in enumerate(species_dict.items(), 2):
        ws_species.cell(row=row_idx, column=1, value=code).font    = Font(name="Arial")
        ws_species.cell(row=row_idx, column=2, value=names[0]).font = Font(name="Arial")
        ws_species.cell(row=row_idx, column=3, value=names[1]).font = Font(name="Arial")

    ws_species.freeze_panes = "A2"
    ws_species.column_dimensions["A"].width = 14
    ws_species.column_dimensions["B"].width = 28
    ws_species.column_dimensions["C"].width = 28

    return wb


def _load_or_create_workbook(
    xlsx_path: pth.Path,
    species_dict: dict,
) -> tuple[openpyxl.Workbook, int]:
    """Return (workbook, next_data_row) for the Tree Labels sheet.

    Loads the existing file when it already exists so that rows can be
    appended without destroying previously written data.
    """
    if xlsx_path.exists():
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["Tree Labels"]
        # Find first truly empty row after the header
        next_row = ws.max_row + 1
        # Guard against trailing blank rows reported by openpyxl
        while next_row > 2 and ws.cell(row=next_row - 1, column=1).value is None:
            next_row -= 1
        return wb, next_row
    else:
        wb = create_workbook(species_dict)
        return wb, 2  # row 1 is the header


def append_tree_rows(
    ws,
    las_idx: int,
    tree_labels: np.ndarray,
    predictions: dict,          # {tree_id: predicted_species_str}  – may be empty
    data_row_start: int,
) -> int:
    """Write one row per tree to *ws* starting at *data_row_start*.

    Column A  – file name  (black)
    Column B  – auto-predicted label  (red = not yet verified by specialist)
    Column C  – verified label  (intentionally left blank for manual entry)

    Returns the next available row index.
    """
    red_font   = Font(name="Arial", color="FF0000")
    black_font = Font(name="Arial")

    for tree_id in sorted(set(tree_labels[tree_labels != -1].tolist())):
        file_name = f"{las_idx}_{tree_id}.npy"
        predicted = predictions.get(int(tree_id), "")

        ws.cell(row=data_row_start, column=1, value=file_name).font = black_font
        cell_pred = ws.cell(row=data_row_start, column=2, value=predicted)
        cell_pred.font = red_font
        # Column C left blank – specialist fills it in
        data_row_start += 1

    return data_row_start


# ──────────────────────────────────────────────────────────────────────────────
# Main pipeline
# ──────────────────────────────────────────────────────────────────────────────

def main(species_dict: dict):
    semantic_labelled_dir = pth.Path("trees/laz")
    laz_paths = list(semantic_labelled_dir.glob("*.laz"))

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    seg = TreeSegmRay(height_min=2.0, max_diameter=0.9, distance_limit=0.3,
                      gravity_factor=0.75, ground_label=1,
                      tree_label=7, verbose=True)
    config_dir = pth.Path(__file__).parent.joinpath("TreePCDClass/src/final_files")
    tree_class_model = LLM_Classifier(resolution=512, species=species_dict,
                                      API_KEY=OPENAI_API_KEY, model=MODEL)

    seg.start_container()

    pbar = tqdm(enumerate(laz_paths), desc = "Processing files", total = len(laz_paths))

    for las_idx, path in pbar:

        dataset_dir = path.parent / path.stem
        dataset_dir.mkdir(exist_ok=True, parents=True)

        xlsx_path = dataset_dir / f"{path.stem}.xlsx"
        pdf_path  = dataset_dir / f"{path.stem}.pdf"
        pdf_name  = f"{path.stem}.pdf"

        # Remove any partial outputs from a previous crashed run so we never
        # append onto an inconsistent file.
        for stale in (xlsx_path, pdf_path):
            if stale.exists():
                stale.unlink()
                # print(f"[CLEANUP] Removed stale {stale.name}")

        try:
            las = laspy.read(path)
            pts = np.vstack([las.x, las.y, las.z]).T.astype(np.float64)
            pts -= pts.mean(axis=0)
            pts = pts.astype(np.float32)
            
            # mask, keep pcd in 50 x 50 m TODO rm later
            # mask = (pts[:, 0] > -10) & (pts[:, 0] < 10) & (pts[:, 1] > -10) & (pts[:, 1] < 10)
            # pts = pts[mask]

            cls = np.asarray(las.classification, dtype=np.int32)
            # cls = cls[mask]

        except Exception as e:

            print(f"Error processing {path}: {e}")
            continue


        try:

            tree_labels = seg.segment(pts, cls)
            tree_xyz    = pts[cls == seg.tree_label]

            valid_mask = tree_labels != -1
            valid_pts  = tree_xyz[valid_mask]
            valid_lbls = tree_labels[valid_mask]

            # Returns {tree_id: raw_prediction_str}; appends pages if PDF exists.
            raw_predictions = save_tree_projections_pdf(
                points=valid_pts,
                labels=valid_lbls,
                las_idx=las_idx,
                source_name=path.stem,
                output_name=pdf_name,
                output_dir=dataset_dir,
                resolution=350,
                device=device,
                tree_classifier=tree_class_model if species_dict is not None else None,
            )

            if species_dict is not None:
                wb, next_row = _load_or_create_workbook(xlsx_path, species_dict)
                append_tree_rows(
                    wb["Tree Labels"],
                    las_idx=las_idx,
                    tree_labels=valid_lbls,
                    predictions=raw_predictions,
                    data_row_start=next_row,
                )
                wb.save(xlsx_path)

        finally:
            seg.rm_container()
            print(f"Prompt tokens used: {tree_class_model.prompt_tokens}",
                  f"Completion tokens used: {tree_class_model.completion_tokens}",
                  f"Cached input tokens: {tree_class_model.cached_tokens}")

if __name__ == "__main__":
    SPECIES = {
            0:  ["Betula_pendula",         "Brzoza brodawkowata"],
            1:  ["Fagus_sylvatica",        "Buk zwyczajny"],
            2:  ["Quercus_petraea",        "Dąb bezszypułkowy"],
            3:  ["Quercus_rubra",          "Dąb czerwony"],
            4:  ["Quercus_robur",          "Dąb szypułkowy"],
            5:  ["Carpinus_betulus",       "Grab pospolity"],
            6:  ["Fraxinus_excelsior",     "Jesion wyniosły"],
            7:  ["Acer_pseudoplatanus",    "Klon jawor"],
            8:  ["Acer_campestre",         "Klon polny"],
            9:  ["Tilia_cordata",          "Lipa drobnolistna"],
            10: ["Ulmus_laevis",           "Wiąz szypułkowy"],
            11: ["Crataegus_monogyna",     "Głóg jednoszyjkowy"],
            12: ["Corylus_avellana",       "Leszczyna pospolita"],
            13: ["Pseudotsuga_menziesii",  "Daglezja zielona"],
            14: ["Abies_alba",             "Jodła pospolita"],
            15: ["Larix_decidua",          "Modrzew europejski"],
            16: ["Pinus_sylvestris",       "Sosna zwyczajna"],
            17: ["Picea_abies",            "Świerk pospolity"],
            18: ["Other",                  "Inne"],
            19: ["Incorrect segmentation", "Błędna segmentacja"],
        }

    main(species_dict=SPECIES)